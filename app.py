# -*- coding: utf-8 -*-
"""
app.py — Sistem Rekapitulasi Absensi (versi dashboard + Supabase)
------------------------------------------------------------------
Login 3 admin (Supabase Auth) -> upload PDF -> data masuk DB sbg batch
draft -> bisa dikoreksi lewat editor tabel -> unduh Excel kapan saja
(tandai "Final" hanya penanda kesiapan, TIDAK mengunci pengeditan,
sesuai keputusan yang diminta).
"""
import os
import io
import functools
import hashlib
import tempfile
import uuid
from datetime import datetime

from flask import (
    Flask, request, jsonify, render_template, redirect,
    url_for, session, send_file
)
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

import db
from extractor import ekstrak_pdf, hitung_signature_pegawai
from rekap_resmi import tulis_sheet_rekap_resmi

app = Flask(__name__)
app.secret_key = os.environ["FLASK_SECRET_KEY"]
app.config["MAX_CONTENT_LENGTH"] = 300 * 1024 * 1024  # 300 MB, cukup utk ratusan PDF


@app.errorhandler(Exception)
def tangani_error_tak_terduga(e):
    """Jaring pengaman terakhir: kalau ada error tak terduga di mana pun,
    tetap balas JSON (bukan halaman HTML traceback bawaan Flask), supaya
    fetch() di dashboard.js tidak gagal parse dan malah menampilkan pesan
    generik "Gagal terhubung ke server" yang membingungkan.

    PERBAIKAN (25 Jul 2026): sebelumnya handler ini menangkap SEMUA
    Exception termasuk HTTPException bawaan Flask (404 Not Found, 405
    Method Not Allowed, dst) - lalu untuk request non-/api/ malah
    di-raise ulang, sehingga 404 biasa (mis. file static/img/logo-
    kejaksaan.png belum diletakkan, atau /favicon.ico yang memang tidak
    ada) tercatat sebagai 500 Internal Server Error di log, padahal itu
    bukan error sungguhan. Sekarang HTTPException dibiarkan lewat apa
    adanya, handler ini hanya aktif untuk exception TAK TERDUGA lainnya."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e  # 404/405/dst biarkan Flask tangani sendiri, apa adanya

    import traceback
    traceback.print_exc()  # tetap tercetak lengkap di terminal untuk ditelusuri
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "pesan": f"Terjadi kesalahan di server: {e}"}), 500
    raise e


# ---------------------------------------------------------------------------
# AUTH
# ---------------------------------------------------------------------------
def login_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "pesan": "Sesi berakhir, silakan login ulang"}), 401
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def master_required(view):
    """Dipasang SETELAH @login_required pada route yang cuma boleh diakses
    Master Admin (manajemen akun). Anggota admin biasa yang mencoba
    memanggil endpoint ini akan ditolak, bukan cuma disembunyikan tombolnya
    di UI - jadi tidak bisa dilewati lewat luar aplikasi."""
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if session.get("user", {}).get("role") != "master":
            return jsonify({"ok": False, "pesan": "Hanya Master Admin yang bisa mengakses fitur ini"}), 403
        return view(*args, **kwargs)
    return wrapped


# ---------------------------------------------------------------------------
# LAPORAN EXCEL RESMI (panel Rincian Kategori di Visualisasi)
# FITUR BARU (14 Agu 2026): dulu file yang diunduh cuma judul + tabel polos
# tanpa identitas maupun styling apa pun. Sekarang dipakai kop institusi +
# tabel bergaris + header berwarna + baris total, konsisten dengan gaya
# rekap resmi (rekap_resmi.py) yang sudah ada di aplikasi ini.
# ---------------------------------------------------------------------------
_WARNA_KOP = "1E3A8A"        # biru tua - nama institusi
_WARNA_HEADER_TABEL = "2563EB"  # biru - header tabel
_WARNA_ZEBRA = "F1F5F9"      # abu sangat muda - baris selang-seling
_WARNA_TOTAL = "DBEAFE"      # biru muda - baris Total Keseluruhan
_GARIS_TIPIS = Side(style="thin", color="CBD5E1")
_BINGKAI_SEL = Border(left=_GARIS_TIPIS, right=_GARIS_TIPIS, top=_GARIS_TIPIS, bottom=_GARIS_TIPIS)


def _kop_laporan_resmi(ws, jumlah_kolom, subjudul, konteks_periode):
    """Kop resmi ala Kejaksaan Tinggi Jawa Tengah - dipakai semua laporan
    Excel yang diunduh dari panel Rincian Kategori di Visualisasi, supaya
    identitasnya konsisten dengan aplikasi, bukan lembar polos tanpa kop.
    Return nomor baris tempat HEADER TABEL sebaiknya mulai ditulis."""
    ws.sheet_view.showGridLines = False
    huruf_akhir = get_column_letter(jumlah_kolom)

    ws.merge_cells(f"A1:{huruf_akhir}1")
    ws["A1"] = "KEJAKSAAN TINGGI JAWA TENGAH — BIDANG DASKRIMTI"
    ws["A1"].font = Font(bold=True, size=13, color=_WARNA_KOP)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{huruf_akhir}2")
    ws["A2"] = subjudul
    ws["A2"].font = Font(bold=True, size=11.5)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A3:{huruf_akhir}3")
    ws["A3"] = konteks_periode
    ws["A3"].font = Font(italic=True, size=9.5, color="64748B")
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A4:{huruf_akhir}4")
    ws["A4"] = f"Dicetak dari Sistem Rekapitulasi Absensi pada {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["A4"].font = Font(italic=True, size=9, color="94A3B8")
    ws["A4"].alignment = Alignment(horizontal="center")

    return 6  # baris 5 sengaja dikosongkan sebagai jarak nafas


def _gaya_header_tabel(ws, baris, jumlah_kolom):
    """Header tabel: teks putih tebal di atas latar biru, rata tengah,
    berbingkai - dipanggil setelah menulis teks header di baris itu."""
    for c in range(1, jumlah_kolom + 1):
        sel = ws.cell(row=baris, column=c)
        sel.font = Font(bold=True, color="FFFFFF", size=10)
        sel.fill = PatternFill("solid", fgColor=_WARNA_HEADER_TABEL)
        sel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sel.border = _BINGKAI_SEL
    ws.row_dimensions[baris].height = 24


def _gaya_baris_data(ws, baris, jumlah_kolom, kolom_rata_tengah, zebra=False):
    """Satu baris data: berbingkai tipis semua sel, kolom yang diminta
    (mis. kolom bulan/angka) dirata-tengah, selebihnya rata kiri, dengan
    warna selang-seling (zebra) supaya mata gampang menyusuri satu baris."""
    for c in range(1, jumlah_kolom + 1):
        sel = ws.cell(row=baris, column=c)
        sel.border = _BINGKAI_SEL
        sel.alignment = Alignment(horizontal="center" if c in kolom_rata_tengah else "left", vertical="center")
        if zebra:
            sel.fill = PatternFill("solid", fgColor=_WARNA_ZEBRA)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", error=None)

    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    user, token, error = db.login(email, password)
    if not user:
        return render_template("login.html", error=error or "Email atau password salah")

    session["user"] = user
    session["token"] = token
    return redirect(url_for("index"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# HALAMAN UTAMA (shell dashboard - semua tab dirender lewat JS/fetch)
# ---------------------------------------------------------------------------
@app.route("/")
@login_required
def index():
    return render_template("dashboard.html", user=session["user"])


# ---------------------------------------------------------------------------
# API: BERANDA (ringkasan angka)
# ---------------------------------------------------------------------------
@app.route("/api/ringkasan-beranda")
@login_required
def ringkasan_beranda():
    # FITUR BARU (1 Agu 2026): "Total batch" & "Perlu ditinjau" tetap GLOBAL
    # (status operasional sistem saat ini, wajar tidak berubah oleh periode)
    # tapi "Pegawai terekap" & kedua ranking Alpha/Terlambat sekarang sama-
    # sama ikut satu filter periode (mode/value dari query string), default
    # ke tahun berjalan seperti sebelumnya. Dulu "Pegawai terekap" dihitung
    # terpisah (jumlah semua batch sepanjang sejarah) sehingga tidak
    # konsisten dengan 2 kartu ranking di bawahnya yang sudah dibatasi
    # tahun berjalan - sekarang ketiganya konsisten pakai statistik_ringkas
    # yang sama, dengan filter yang sama.
    mode = request.args.get("filter_mode", "tahun")
    value = request.args.get("filter_value") or str(datetime.now().year)

    semua = db.daftar_batch()
    total_batch = len(semua)
    batch_draft = [b for b in semua if b["status"] == "draft"]
    perlu_ditinjau = len(batch_draft)
    aktivitas = db.log_aktivitas(limit=8)
    statistik = db.statistik_ringkas(mode, value)

    return jsonify({
        "total_batch": total_batch,
        "perlu_ditinjau": perlu_ditinjau,
        # FITUR BARU (19 Agu 2026): daftar RIL batch berstatus Draft (bukan
        # cuma angkanya) - dibatasi 6 batch terbaru - supaya Beranda bisa
        # menampilkan pintasan langsung "buka batch ini" alih-alih admin
        # harus pindah ke tab Riwayat dulu untuk mencari sendiri batch mana
        # saja yang masih Draft.
        "batch_draft": batch_draft[:6],
        "total_pegawai": statistik["total_pegawai"],
        "pegawai_unik": statistik["pegawai_unik"],
        "aktivitas_terbaru": aktivitas,
        "ranking_alpha": db.ranking_pegawai("alpha", mode, value, limit=15),
        "ranking_terlambat": db.ranking_pegawai("terlambat", mode, value, limit=15),
        # FITUR BARU (19 Agu 2026): tren bulanan (sama seperti grafik garis
        # di tab Visualisasi) ikut dikirim ke Beranda supaya bisa ditampilkan
        # sebagai preview kecil + dipakai menghitung naik/turunnya Alpha
        # dibanding bulan sebelumnya.
        "tren": db.tren_bulanan(mode, value),
    })


# ---------------------------------------------------------------------------
# API: PROSES BATCH BARU — dipecah jadi 3 langkah (mulai -> per file -> selesai)
# supaya browser bisa menampilkan progres ASLI (bukan animasi palsu), karena
# tiap file diproses lewat request terpisah dan hasilnya langsung diketahui.
# ---------------------------------------------------------------------------
def _hash_file(path):
    """SHA-256 dari isi file (byte-per-byte) - dipakai deteksi duplikat
    lapis 1 di proses_satu_file(). File dengan isi identik akan punya hash
    yang sama persis walau nama filenya berbeda."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for potongan in iter(lambda: fh.read(8192), b""):
            h.update(potongan)
    return h.hexdigest()


@app.route("/api/proses/mulai", methods=["POST"])
@login_required
def proses_mulai():
    nama_bidang = request.form.get("nama_bidang", "").strip()
    label = nama_bidang.title() if nama_bidang else f"Batch {datetime.now().strftime('%d %b %Y %H:%M')}"
    batch = db.buat_batch(
        nama_bidang=nama_bidang,
        label=label,
        jumlah_pegawai=0,
        dibuat_oleh=session["user"]["email"],
    )
    return jsonify({"ok": True, "batch_id": batch["id"]})


@app.route("/api/proses/file", methods=["POST"])
@login_required
def proses_satu_file():
    batch_id = request.form.get("batch_id")
    f = request.files.get("file")
    if not batch_id or not f:
        return jsonify({"ok": False, "pesan": "batch_id atau file tidak ada"}), 400

    ekstensi = os.path.splitext(f.filename)[1] or ".pdf"
    tmp_path = os.path.join(tempfile.gettempdir(), f"absensi_{uuid.uuid4().hex}{ekstensi}")
    f.save(tmp_path)

    # --- DETEKSI DUPLIKAT LAPIS 1: hash isi file (byte-per-byte) ---
    # Menangkap kasus file yang sama persis terunggah dua kali (nama file
    # boleh beda, isinya yang dibandingkan). Dicek SEBELUM file diekstrak
    # sama sekali, supaya tidak ada kerja/parsing yang sia-sia.
    file_hash = _hash_file(tmp_path)
    file_asli_sama = db.cek_dan_catat_hash_file(batch_id, file_hash, f.filename)
    if file_asli_sama:
        os.remove(tmp_path)
        label_asal = file_asli_sama.get("batch_label") or ""
        info_batch = f" (di batch '{label_asal}')" if label_asal else ""
        alasan = (f"File ini identik (isi byte sama persis) dengan '{file_asli_sama['nama_file']}'"
                  f"{info_batch} yang sudah pernah diproses sebelumnya - kemungkinan besar "
                  f"terunggah dua kali. Data dari file ini TIDAK disimpan.")
        db.simpan_hasil_ekstraksi(batch_id, [], [], [(f.filename, alasan)])
        return jsonify({
            "ok": True, "nama_file": f.filename, "bermasalah": True,
            "alasan": alasan, "jumlah_pegawai_baru": 0,
        })

    rows, ringkasan, error = [], [], None
    try:
        rows, ringkasan, error = ekstrak_pdf(tmp_path, f.filename)
    except Exception as e:
        error = f"Gagal diproses: {e}"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    # Bidang manual (diisi di form Proses Batch Baru) berlaku untuk SELURUH
    # pegawai di file ini. Kalau dikosongkan (upload gabungan lintas Bidang),
    # Bidang tiap pegawai dibiarkan kosong ("Belum diketahui") dan bisa
    # dikoreksi satu per satu nanti di tab Ringkasan Pegawai.
    nama_bidang_manual = request.form.get("nama_bidang", "").strip()
    if nama_bidang_manual and not error:
        for r in rows:
            r["Bidang"] = nama_bidang_manual.upper()
        for r in ringkasan:
            r["Bidang"] = nama_bidang_manual.upper()

    peringatan_duplikat = []
    if not error and rows:
        # --- DETEKSI DUPLIKAT LAPIS 2: "sidik jari" data harian per pegawai ---
        # Menangkap kasus file DIEKSPOR ULANG dari sistem sumber - isi data
        # per pegawai sama persis, tapi byte file berbeda (mis. metadata/
        # timestamp cetak ulang PDF-nya beda) sehingga tidak tertangkap hash
        # lapis 1 di atas. Dicek per-pegawai, bukan per-file, supaya kalau
        # cuma SEBAGIAN pegawai di file ini yang ternyata duplikat, pegawai
        # lain yang datanya baru tetap tersimpan seperti biasa.
        per_nip = {}
        for r in rows:
            per_nip.setdefault(r["NIP"], []).append(r)

        nip_duplikat = set()
        for nip, baris_pegawai in per_nip.items():
            if nip == "-":
                continue  # identitas tidak jelas, tidak bisa dipakai jadi kunci pembanding
            sig = hitung_signature_pegawai(baris_pegawai)
            file_asli = db.cek_dan_catat_signature_pegawai(batch_id, nip, sig, f.filename)
            if file_asli:
                nip_duplikat.add(nip)
                nama = baris_pegawai[0]["Nama"]
                label_asal = file_asli.get("batch_label") or ""
                info_batch = f" (di batch '{label_asal}')" if label_asal else ""
                peringatan_duplikat.append(
                    f"{nama} (NIP {nip}) dilewati dari file ini - data hariannya identik "
                    f"dengan yang sudah tersimpan dari file '{file_asli['nama_file']}'{info_batch} "
                    f"(indikasi file diekspor ulang dari sistem sumber / terunggah dua kali)."
                )

        if nip_duplikat:
            rows = [r for r in rows if r["NIP"] not in nip_duplikat]
            ringkasan = [r for r in ringkasan if r["NIP"] not in nip_duplikat]

    berkas_bermasalah = []
    if error:
        berkas_bermasalah.append((f.filename, error))
    if peringatan_duplikat:
        berkas_bermasalah.append((f.filename, " | ".join(peringatan_duplikat)))

    try:
        db.simpan_hasil_ekstraksi(batch_id, rows if not error else [], ringkasan if not error else [], berkas_bermasalah)
    except Exception as e:
        return jsonify({"ok": False, "pesan": f"Gagal menyimpan {f.filename} ke database: {e}"}), 500

    return jsonify({
        "ok": True,
        "nama_file": f.filename,
        "bermasalah": bool(error) or bool(peringatan_duplikat),
        "alasan": error or (" | ".join(peringatan_duplikat) if peringatan_duplikat else None),
        "jumlah_pegawai_baru": len(ringkasan) if not error else 0,
    })


@app.route("/api/proses/selesai", methods=["POST"])
@login_required
def proses_selesai():
    data = request.get_json(force=True)
    batch_id = data.get("batch_id")
    jumlah_pegawai = db.perbarui_jumlah_pegawai(batch_id)
    periode_awal, periode_akhir = db.perbarui_periode_batch(batch_id)
    attendance = db.ambil_attendance(batch_id)
    log_bermasalah = db.ambil_berkas_bermasalah(batch_id)

    # FITUR BARU (1 Agu 2026): dulu Log Aktivitas cuma mencatat edit field
    # satu-satu, jadi upload batch (aktivitas paling sering terjadi) sama
    # sekali tidak tercatat. Sekarang dicatat juga, dengan ringkasan jumlah
    # file & pegawai supaya langsung informatif tanpa perlu buka detail batch.
    batch = db.ambil_batch(batch_id)
    jumlah_bermasalah = len(log_bermasalah)
    detail = f"{jumlah_pegawai} pegawai" + (f", {jumlah_bermasalah} berkas bermasalah" if jumlah_bermasalah else "")
    db.catat_log_batch(batch_id, batch.get("label", ""), "Upload Batch", detail, session["user"]["email"])

    return jsonify({
        "ok": True,
        "batch_id": batch_id,
        "jumlah_pegawai": jumlah_pegawai,
        "periode_awal": periode_awal,
        "periode_akhir": periode_akhir,
        "jumlah_baris": len(attendance),
        "pratinjau": attendance[:10],
        "log_bermasalah": [{"nama_file": e["nama_file"], "alasan": e["alasan"]} for e in log_bermasalah],
    })


# ---------------------------------------------------------------------------
# API: RIWAYAT BATCH (daftar + detail + edit + unduh + hapus)
# ---------------------------------------------------------------------------
@app.route("/api/batches")
@login_required
def api_daftar_batch():
    bidang = request.args.get("bidang") or None
    status = request.args.get("status") or None
    return jsonify(db.daftar_batch(bidang, status))


@app.route("/api/tahun-tersedia")
@login_required
def api_tahun_tersedia():
    return jsonify(db.daftar_tahun_tersedia())


@app.route("/api/batches/<batch_id>")
@login_required
def api_detail_batch(batch_id):
    return jsonify({
        "batch": db.ambil_batch(batch_id),
        "attendance": db.ambil_attendance(batch_id),
        "ringkasan": db.ambil_ringkasan(batch_id),
        "berkas_bermasalah": db.ambil_berkas_bermasalah(batch_id),
    })


@app.route("/api/batches/<batch_id>/final", methods=["POST"])
@login_required
def api_tandai_final(batch_id):
    db.tandai_final(batch_id)
    return jsonify({"ok": True})


@app.route("/api/batches/<batch_id>/draf", methods=["POST"])
@login_required
def api_tandai_draf(batch_id):
    db.tandai_draft(batch_id)
    return jsonify({"ok": True})


@app.route("/api/batches/<batch_id>", methods=["DELETE"])
@login_required
def api_hapus_batch(batch_id):
    # Catat log SEBELUM batch benar-benar dihapus (perlu label & jumlah
    # pegawainya dulu selagi masih ada) - lihat catatan "on delete set null"
    # di schema.sql supaya baris log ini sendiri tidak ikut terhapus.
    batch = db.ambil_batch(batch_id)
    if batch:
        detail = f"{batch.get('jumlah_pegawai', 0)} pegawai, periode {batch.get('periode_awal') or '-'} s.d. {batch.get('periode_akhir') or '-'}"
        db.catat_log_batch(batch_id, batch.get("label", ""), "Hapus Batch", detail, session["user"]["email"])
    db.hapus_batch(batch_id)
    return jsonify({"ok": True})


@app.route("/api/edit", methods=["POST"])
@login_required
def api_edit():
    """Body JSON: {batch_id, perubahan: [{record_table, record_id, field, nilai_baru}, ...]}"""
    data = request.get_json(force=True)
    batch_id = data["batch_id"]

    # Proteksi di server (bukan cuma UI): batch yang sudah final tidak boleh
    # diedit sampai dikembalikan ke draf dulu, supaya rekap yang sudah
    # difinalisasi/diunduh tidak diam-diam berubah tanpa disadari.
    batch = db.ambil_batch(batch_id)
    if batch.get("status") == "final":
        return jsonify({
            "ok": False,
            "pesan": "Batch ini berstatus Final dan tidak bisa diedit. Klik \"Tandai Draf lagi\" dulu di halaman batch untuk membuka kembali.",
        }), 409

    hasil = []
    for p in data.get("perubahan", []):
        hasil.append(db.edit_field(batch_id, p["record_table"], p["record_id"], p["field"], p["nilai_baru"], session["user"]["email"]))
    # "hasil" membawa balik nilai TERBARU tiap baris yang diedit (termasuk
    # kolom Telat & ringkasan_pegawai yang ikut dihitung ulang otomatis -
    # lihat db.py::edit_field) supaya frontend bisa merefresh tampilan tanpa
    # perlu reload seluruh batch. Begitu perubahan ini disimpan, baris
    # dianggap langsung final - tidak ada langkah konfirmasi terpisah.
    return jsonify({"ok": True, "hasil": hasil})


@app.route("/api/batches/<batch_id>/unduh")
@login_required
def api_unduh(batch_id):
    batch = db.ambil_batch(batch_id)

    # Proteksi di server: hanya batch berstatus Final yang boleh diunduh,
    # supaya rekap yang terunduh selalu sudah "dikunci"/disetujui, bukan
    # data draf yang masih mungkin berubah.
    if batch.get("status") != "final":
        return jsonify({
            "ok": False,
            "pesan": "Batch ini masih berstatus Draf. Tandai Final dulu sebelum bisa diunduh.",
        }), 409

    ringkasan_db = db.ambil_ringkasan(batch_id)

    # PERBAIKAN (23 Jul 2026): periode diambil langsung dari kolom
    # batches.periode_awal/periode_akhir (sudah dihitung & disimpan saat
    # batch selesai diproses - lihat db.py::perbarui_periode_batch),
    # bukan dihitung ulang dari attendance_records tiap kali unduh. Ini
    # memperbaiki periode yang tampil "-" s.d. "-" di Excel (akibat format
    # tanggal lama vs ISO yang tidak sinkron), dan sekaligus menghindari
    # perlu ambil ribuan baris attendance_records hanya untuk cari
    # tanggal awal/akhir.
    periode_awal = batch.get("periode_awal")
    periode_akhir = batch.get("periode_akhir")
    if not periode_awal or not periode_akhir:
        # jaga-jaga: batch lama yang belum sempat terisi periodenya
        periode_awal, periode_akhir = db.perbarui_periode_batch(batch_id)

    # kembalikan ke bentuk dict yang dipakai tulis_sheet_rekap_resmi (rekap_resmi.py)
    ringkasan_list = []
    for r in ringkasan_db:
        ringkasan_list.append({
            "Nama": r["nama"], "NIP": r["nip"], "NRP": r["nrp"], "Golongan": r["golongan"],
            "Terlambat (Hari)": r["terlambat"], "Pulang Cepat (Hari)": r["pulang_cepat"],
            "Tidak Absen Datang (Hari)": r["tidak_absen_datang"],
            "Tidak Absen Pulang (Hari)": r["tidak_absen_pulang"],
            "Izin (Hari)": r["izin"], "Alpha (Hari)": r["alpha"], "Sakit (Hari)": r["sakit"],
            "Dinas Luar (Hari)": r["dinas_luar"], "Lepas Piket (Hari)": r["lepas_piket"],
            "Tugas Belajar (Hari)": r["tugas_belajar"], "Total Cuti (Hari)": r["total_cuti"],
            "Rincian Cuti": r["rincian_cuti"], "Total Hari Kerja": r["total_hari_kerja"],
            "Sumber File": r["sumber_file"],
            "_sub_unit": r.get("sub_unit_kerja", ""), "_jabatan": r.get("jabatan", ""),
        })

    wb = Workbook()
    wb.remove(wb.active)
    tulis_sheet_rekap_resmi(wb, ringkasan_list, periode_awal, periode_akhir, batch.get("nama_bidang", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    db.catat_unduhan(batch_id, session["user"]["email"])
    nama_file = f"Rekap_{batch.get('label', 'Batch').replace(' ', '_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nama_file,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# API: CARI PEGAWAI
# ---------------------------------------------------------------------------
@app.route("/api/visualisasi")
@login_required
def visualisasi():
    mode = request.args.get("filter_mode", "all")
    value = request.args.get("filter_value")
    return jsonify({
        "statistik": db.statistik_ringkas(mode, value),
        "keterangan": db.agregasi_keterangan(mode, value),
        # PERBAIKAN (30 Jul 2026): dulu selalu "all" (tidak ikut filter,
        # jadi acuan pembanding tetap). Sekarang Filter Bidang sudah
        # dihapus dari halaman ini dan cuma tersisa Filter Batch - jadi
        # Tren Bulanan ikut mengikuti batch yang dipilih juga.
        "tren": db.tren_bulanan(mode, value),
        "ranking_alpha": db.ranking_pegawai("alpha", mode, value, limit=15),
        # PERBAIKAN (30 Jul 2026): tadinya 10, dinaikkan jadi 15 supaya
        # grafik batang Rekapitulasi Keterlambatan bisa memuat lebih banyak
        # pegawai sekaligus.
        "ranking_terlambat": db.ranking_pegawai("terlambat", mode, value, limit=15),
    })


# Kolom ringkasan_pegawai yang boleh diminta lewat kartu statistik yang
# bisa diklik - whitelist ini SENGAJA dijaga ketat (bukan menerima nama
# field bebas dari query string) karena nilainya ditempel langsung ke
# string SELECT Supabase (lihat db.py::ranking_pegawai).
_FIELD_STAT_DIIZINKAN = {"terlambat", "sakit", "izin", "alpha"}

# Daftar kategori Keterangan resmi - HARUS selalu sama persis dengan daftar
# di dashboard.js::WARNA_KETERANGAN (dipakai buat legenda donat), supaya
# kategori "Lainnya" (label yang TIDAK ada di daftar ini) konsisten antara
# apa yang ditampilkan di donat dan apa yang muncul saat diklik.
_KATEGORI_KETERANGAN_RESMI = [
    "WFO", "Libur", "Izin", "Dinas Luar", "Alpha", "Sakit",
    "Cuti", "Cuti Alasan Penting", "Cuti Besar", "Cuti Belajar", "Lepas Piket",
]


@app.route("/api/visualisasi/rincian")
@login_required
def api_visualisasi_rincian():
    """FITUR BARU (11 Agu 2026): dipanggil saat kartu statistik (Telat/
    Sakit/Izin/Alpha) ATAU baris legenda donat Komposisi Keterangan di
    Visualisasi diklik - mengembalikan daftar LENGKAP semua pegawai untuk
    kategori itu (bukan cuma top-N ringkasan singkat), supaya bisa
    ditampilkan sebagai tabel rincian yang bisa diurutkan nama/jumlah."""
    mode = request.args.get("filter_mode", "all")
    value = request.args.get("filter_value")
    jenis = request.args.get("jenis")  # 'stat' | 'keterangan' | 'lainnya'
    if jenis == "stat":
        field = request.args.get("field", "")
        if field not in _FIELD_STAT_DIIZINKAN:
            return jsonify({"ok": False, "pesan": "Field tidak dikenal."}), 400
        return jsonify(db.ranking_pegawai(field, mode, value, limit=None))
    if jenis == "keterangan":
        label = request.args.get("label", "")
        return jsonify(db.ranking_keterangan_harian(label, mode, value))
    if jenis == "lainnya":
        return jsonify(db.ranking_lainnya(mode, value, _KATEGORI_KETERANGAN_RESMI))
    return jsonify({"ok": False, "pesan": "Jenis rincian tidak dikenal."}), 400


@app.route("/api/visualisasi/rincian-bulanan")
@login_required
def api_visualisasi_rincian_bulanan():
    """FITUR BARU (13 Agu 2026): versi "tabel bulanan" panel rincian
    kategori - per pegawai, dipecah Jan..Des untuk satu tahun tertentu +
    kolom Total. Beda dari api_visualisasi_rincian yang cuma satu angka
    total per pegawai (tidak dipecah per bulan)."""
    jenis = request.args.get("jenis")  # 'stat' | 'keterangan'
    tahun = request.args.get("tahun") or str(datetime.now().year)
    if jenis == "stat":
        field = request.args.get("field", "")
        if field not in _FIELD_STAT_DIIZINKAN:
            return jsonify({"ok": False, "pesan": "Field tidak dikenal."}), 400
        return jsonify({"tahun": tahun, "bulan_label": db.BULAN_LABEL, "data": db.rincian_bulanan("stat", field, tahun)})
    if jenis == "keterangan":
        label = request.args.get("label", "")
        return jsonify({"tahun": tahun, "bulan_label": db.BULAN_LABEL, "data": db.rincian_bulanan("keterangan", label, tahun)})
    return jsonify({"ok": False, "pesan": "Jenis rincian tidak dikenal."}), 400


@app.route("/api/visualisasi/rincian/unduh")
@login_required
def api_visualisasi_rincian_unduh():
    """Unduh rincian yang sama (lihat api_visualisasi_rincian) sebagai
    laporan Excel - dipakai tombol "Unduh sebagai laporan" di panel
    rincian kategori Visualisasi."""
    mode = request.args.get("filter_mode", "all")
    value = request.args.get("filter_value")
    jenis = request.args.get("jenis")
    urut = request.args.get("urut", "jumlah")  # 'jumlah' | 'nama'
    kolom_label_asli = False

    if jenis == "stat":
        field = request.args.get("field", "")
        if field not in _FIELD_STAT_DIIZINKAN:
            return jsonify({"ok": False, "pesan": "Field tidak dikenal."}), 400
        data = db.ranking_pegawai(field, mode, value, limit=None)
        judul = field.capitalize()
    elif jenis == "keterangan":
        judul = request.args.get("label", "")
        data = db.ranking_keterangan_harian(judul, mode, value)
    elif jenis == "lainnya":
        judul = "Lainnya"
        data = db.ranking_lainnya(mode, value, _KATEGORI_KETERANGAN_RESMI)
        kolom_label_asli = True
    else:
        return jsonify({"ok": False, "pesan": "Jenis rincian tidak dikenal."}), 400

    if urut == "nama":
        data = sorted(data, key=lambda x: (x.get("nama") or "").lower())

    label_periode = value if mode in ("tahun", "batch") and value else "Seluruh Batch (akumulasi total)"
    jumlah_kolom = 5 if kolom_label_asli else 4
    wb = Workbook()
    ws = wb.active
    ws.title = "Rincian"

    baris = _kop_laporan_resmi(ws, jumlah_kolom, f"Rincian {judul}", f"{label_periode} · {len(data)} baris tercatat")

    header = ["No", "Nama", "NIP", "Label Asli", "Jumlah Hari"] if kolom_label_asli else ["No", "Nama", "NIP", "Jumlah Hari"]
    for c, teks in enumerate(header, 1):
        ws.cell(row=baris, column=c, value=teks)
    _gaya_header_tabel(ws, baris, jumlah_kolom)
    baris += 1

    kolom_tengah = {1, jumlah_kolom}  # No + Jumlah Hari rata tengah, sisanya rata kiri
    for i, d in enumerate(data):
        nilai_baris = [i + 1, d.get("nama") or "-", d.get("nip") or "-"]
        if kolom_label_asli:
            nilai_baris.append(d.get("label_asli") or "-")
        nilai_baris.append(d.get("jumlah") or 0)
        for c, v in enumerate(nilai_baris, 1):
            ws.cell(row=baris, column=c, value=v)
        _gaya_baris_data(ws, baris, jumlah_kolom, kolom_tengah, zebra=(i % 2 == 1))
        baris += 1

    # Baris "Total Keseluruhan" - jumlah kolom Jumlah Hari, supaya laporan
    # langsung kelihatan akumulasinya tanpa perlu dijumlah manual.
    ws.cell(row=baris, column=2, value="TOTAL KESELURUHAN")
    ws.merge_cells(start_row=baris, start_column=2, end_row=baris, end_column=jumlah_kolom - 1)
    ws.cell(row=baris, column=jumlah_kolom, value=sum(d.get("jumlah") or 0 for d in data))
    for c in range(1, jumlah_kolom + 1):
        sel = ws.cell(row=baris, column=c)
        sel.font = Font(bold=True)
        sel.fill = PatternFill("solid", fgColor=_WARNA_TOTAL)
        sel.border = _BINGKAI_SEL
        sel.alignment = Alignment(horizontal="center" if c != 2 else "right", vertical="center")

    lebar = (5, 34, 22, 20, 12) if kolom_label_asli else (5, 34, 22, 12)
    for idx, w in enumerate(lebar, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"rincian_{judul.replace(' ', '_')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/visualisasi/rincian-bulanan/unduh")
@login_required
def api_visualisasi_rincian_bulanan_unduh():
    """Unduh versi bulanan (lihat api_visualisasi_rincian_bulanan) sebagai
    laporan Excel - kolom Nama/NIP + Jan..Des + Total."""
    jenis = request.args.get("jenis")
    tahun = request.args.get("tahun") or str(datetime.now().year)
    urut = request.args.get("urut", "jumlah")  # 'jumlah' (=Total) | 'nama'

    if jenis == "stat":
        field = request.args.get("field", "")
        if field not in _FIELD_STAT_DIIZINKAN:
            return jsonify({"ok": False, "pesan": "Field tidak dikenal."}), 400
        judul = field.capitalize()
        data = db.rincian_bulanan("stat", field, tahun)
    elif jenis == "keterangan":
        judul = request.args.get("label", "")
        data = db.rincian_bulanan("keterangan", judul, tahun)
    else:
        return jsonify({"ok": False, "pesan": "Jenis rincian tidak dikenal."}), 400

    if urut == "nama":
        data = sorted(data, key=lambda x: (x.get("nama") or "").lower())

    jumlah_kolom = 3 + len(db.BULAN_LABEL) + 1  # No, Nama, NIP, 12 bulan, Total
    wb = Workbook()
    ws = wb.active
    ws.title = "Rincian Bulanan"

    baris = _kop_laporan_resmi(
        ws, jumlah_kolom,
        f"Rincian {judul} per Bulan",
        f"Tahun {tahun} · {len(data)} pegawai tercatat",
    )

    header = ["No", "Nama", "NIP"] + db.BULAN_LABEL + ["Total"]
    for c, teks in enumerate(header, 1):
        ws.cell(row=baris, column=c, value=teks)
    _gaya_header_tabel(ws, baris, jumlah_kolom)
    baris += 1

    kolom_tengah = {1} | set(range(4, jumlah_kolom + 1))  # No + semua kolom bulan + Total
    for i, d in enumerate(data):
        nilai_baris = [i + 1, d.get("nama") or "-", d.get("nip") or "-"] + [d.get(b, 0) for b in db.BULAN_URUT] + [d.get("total", 0)]
        for c, v in enumerate(nilai_baris, 1):
            ws.cell(row=baris, column=c, value=v)
        _gaya_baris_data(ws, baris, jumlah_kolom, kolom_tengah, zebra=(i % 2 == 1))
        baris += 1

    # Baris "Total Keseluruhan" - jumlah tiap kolom bulan + grand total,
    # supaya laporan langsung kelihatan akumulasi seluruh pegawai tanpa
    # perlu dijumlah manual di Excel.
    ws.cell(row=baris, column=2, value="TOTAL KESELURUHAN")
    ws.merge_cells(start_row=baris, start_column=2, end_row=baris, end_column=3)
    for idx, b in enumerate(db.BULAN_URUT):
        ws.cell(row=baris, column=4 + idx, value=sum(d.get(b, 0) for d in data))
    ws.cell(row=baris, column=jumlah_kolom, value=sum(d.get("total", 0) for d in data))
    for c in range(1, jumlah_kolom + 1):
        sel = ws.cell(row=baris, column=c)
        sel.font = Font(bold=True)
        sel.fill = PatternFill("solid", fgColor=_WARNA_TOTAL)
        sel.border = _BINGKAI_SEL
        sel.alignment = Alignment(horizontal="center" if c != 2 else "right", vertical="center")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    for idx in range(len(db.BULAN_LABEL)):
        ws.column_dimensions[get_column_letter(4 + idx)].width = 6.5
    ws.column_dimensions[get_column_letter(jumlah_kolom)].width = 9

    # Bekukan panel: baris kop+header tidak ikut tergulung, kolom Nama/NIP
    # tetap terlihat saat digeser ke kolom bulan yang jauh di kanan.
    ws.freeze_panes = ws.cell(row=baris - len(data), column=4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"rincian_{judul.replace(' ', '_')}_bulanan_{tahun}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/cari-pegawai")
@login_required
def api_cari_pegawai():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    return jsonify(db.cari_pegawai(q))


@app.route("/api/riwayat-pegawai/<nip>")
@login_required
def api_riwayat_pegawai(nip):
    return jsonify(db.riwayat_pegawai(nip))


# ---------------------------------------------------------------------------
# API: LOG AKTIVITAS (global, lintas batch)
# ---------------------------------------------------------------------------
@app.route("/api/log-aktivitas")
@login_required
def api_log_aktivitas():
    limit = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0, type=int)
    return jsonify(db.log_aktivitas(limit=limit, offset=offset))


# ---------------------------------------------------------------------------
# API: PENGATURAN — daftar keterangan (dropdown)
# ---------------------------------------------------------------------------
@app.route("/api/keterangan", methods=["GET", "POST"])
@login_required
def api_keterangan():
    if request.method == "POST":
        label = request.get_json(force=True).get("label", "").strip()
        if label:
            db.tambah_keterangan(label)
        return jsonify({"ok": True})
    return jsonify(db.daftar_keterangan())


@app.route("/api/keterangan/<int:keterangan_id>", methods=["DELETE"])
@login_required
def api_hapus_keterangan(keterangan_id):
    db.hapus_keterangan(keterangan_id)
    return jsonify({"ok": True})


@app.route("/api/bidang", methods=["GET", "POST"])
@login_required
def api_bidang():
    if request.method == "POST":
        label = request.get_json(force=True).get("label", "").strip()
        if label:
            db.tambah_bidang(label)
        return jsonify({"ok": True})
    return jsonify(db.daftar_bidang())


@app.route("/api/pengaturan/<kunci>", methods=["GET", "PUT"])
@login_required
def api_pengaturan(kunci):
    if request.method == "PUT":
        nilai = request.get_json(force=True).get("nilai", "")
        db.set_pengaturan(kunci, nilai)
        return jsonify({"ok": True})
    return jsonify({"kunci": kunci, "nilai": db.ambil_pengaturan(kunci)})


@app.route("/api/sinkronkan-ringkasan", methods=["POST"])
@login_required
def api_sinkronkan_ringkasan():
    """FITUR BARU (13 Agu 2026): tombol "Sinkronkan Ulang Ringkasan" di
    halaman Pengaturan - menghitung ulang statistik ringkasan_pegawai
    (Alpha, Sakit, Izin, Terlambat, dst) untuk SEMUA pegawai/batch
    sekaligus dari data harian saat ini, supaya sama persis dengan
    hitungan yang dipakai donut chart Visualisasi. Bisa memakan waktu
    untuk data yang sangat banyak (satu query per pegawai per batch)."""
    jumlah = db.sinkronkan_semua_ringkasan()
    return jsonify({"ok": True, "jumlah": jumlah})


@app.route("/api/bidang/<int:bidang_id>", methods=["DELETE"])
@login_required
def api_hapus_bidang(bidang_id):
    db.hapus_bidang(bidang_id)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# AKUN (khusus Master Admin) - tambah/nonaktifkan/hapus admin lain lewat
# dashboard, tanpa perlu buka Supabase Dashboard langsung.
# ---------------------------------------------------------------------------
@app.route("/api/akun", methods=["GET", "POST"])
@login_required
@master_required
def api_akun():
    if request.method == "POST":
        data = request.get_json(force=True)
        email = data.get("email", "").strip()
        password = data.get("password", "")
        nama = data.get("nama", "").strip()
        if not email or not password:
            return jsonify({"ok": False, "pesan": "Email dan password wajib diisi"}), 400
        if len(password) < 6:
            return jsonify({"ok": False, "pesan": "Password minimal 6 karakter"}), 400
        try:
            db.buat_akun_admin(email, password, nama)
        except Exception as e:
            return jsonify({"ok": False, "pesan": f"Gagal membuat akun: {e}"}), 400
        return jsonify({"ok": True})
    return jsonify(db.daftar_akun())


@app.route("/api/akun/<user_id>/status", methods=["POST"])
@login_required
@master_required
def api_status_akun(user_id):
    if user_id == session["user"]["id"]:
        return jsonify({"ok": False, "pesan": "Tidak bisa menonaktifkan akun sendiri"}), 400
    aktif = bool(request.get_json(force=True).get("aktif", True))
    db.set_status_akun(user_id, aktif)
    return jsonify({"ok": True})


@app.route("/api/akun/<user_id>", methods=["DELETE"])
@login_required
@master_required
def api_hapus_akun(user_id):
    if user_id == session["user"]["id"]:
        return jsonify({"ok": False, "pesan": "Tidak bisa menghapus akun sendiri"}), 400
    profil = db.ambil_profil(user_id)
    if profil and profil.get("role") == "master":
        return jsonify({"ok": False, "pesan": "Tidak bisa menghapus akun Master Admin lewat sini"}), 400
    try:
        db.hapus_akun(user_id)
    except Exception as e:
        return jsonify({"ok": False, "pesan": f"Gagal menghapus akun: {e}"}), 400
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
